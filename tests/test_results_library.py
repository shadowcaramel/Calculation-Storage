"""
Tests for the results-library reader: catalog, annotations, migrations, lint, views, CLI.

Run with:
    pytest tests/test_results_library.py
"""

from __future__ import annotations

import json
import sqlite3
import sys
from pathlib import Path

import pandas as pd
import pytest

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from results_library.annotations import (  # noqa: E402
    lint_annotations,
    load_annotations,
    merge_annotation,
)
from results_library.catalog import (  # noqa: E402
    build_catalog,
    load_catalog,
    scan_bundles,
    write_catalog,
)
from results_library.cli import main as cli_main  # noqa: E402
from results_library.lint import lint_library  # noqa: E402
from results_library.migrations import apply_migrations, latest_version  # noqa: E402
from results_library.sync_schema import (  # noqa: E402
    FILES as SCHEMA_FILES,
    source_dir,
    sync as sync_schema,
)
from results_library.views.latex import csv_table, latex_table, tsv_table  # noqa: E402
from results_schema.models import (  # noqa: E402
    Computed,
    Observable,
    Provenance,
    Reference,
    ResultRecord,
    StateRecord,
    StoredFile,
    Subject,
    Variant,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_PLOT = b"fake-png-bytes"
_CSV = "x,y\n1,2\n"


def _computed(**overrides) -> Computed:
    base = dict(
        median=2.639,
        Q1=2.637,
        Q3=2.641,
        err_low=0.002,
        err_high=0.002,
        IQR=0.004,
        N_models=12,
        Nmax_final=8,
        uncertainty_method="IQR",
        homega_aggregation="min",
    )
    base.update(overrides)
    return Computed(**base)


def _he6_record(run_stamp: str, **kwargs) -> ResultRecord:
    artifacts = kwargs.pop(
        "artifacts",
        {
            "histogram_plot": "histogram.png",
            "histogram_data": "histogram.parquet",
            "histogram_data_csv": "histogram.csv",
        },
    )
    available = kwargs.pop("available", ["histogram"])
    return ResultRecord.build(
        subject=Subject.single(
            "6He", StateRecord(J2=0, parity="+", T2=2, index=2)
        ),
        observable=Observable(
            slug="Erel", name="Erel", latex="$E_{rel}$", units="MeV"
        ),
        computed=kwargs.pop("computed", _computed()),
        run_stamp=run_stamp,
        variant=kwargs.pop(
            "variant",
            Variant(
                selection_set="final",
                bounds={"Nmax": [2, 6], "hOmega": [13, 40]},
                potential="Daejeon16",
            ),
        ),
        provenance=kwargs.pop(
            "provenance",
            Provenance(source_workbook="6He.xlsx", source_sheet="0p2"),
        ),
        available=available,
        artifacts=artifacts,
        **kwargs,
    )


def _c16_transition(run_stamp: str = "2026-08-18_be2") -> ResultRecord:
    initial = StateRecord(J2=4, parity="+", T2=4, index=2)
    final = StateRecord(J2=4, parity="+", T2=4, index=1)
    return ResultRecord.build(
        subject=Subject.transition("16C", initial, final),
        observable=Observable(
            slug="BE2", name="B(E2)", latex="B(E2)", units="e^2 fm^4", direction="down"
        ),
        computed=_computed(median=8.492, err_low=0.749, err_high=0.543, N_models=261),
        run_stamp=run_stamp,
        artifacts={"histogram_plot": "histogram.png"},
        available=["histogram"],
    )


def _c17_relative(run_stamp: str = "2026-08-18_erel") -> ResultRecord:
    return ResultRecord.build(
        subject=Subject.single(
            "17C", StateRecord(J2=5, parity="+", T2=5, index=1)
        ),
        observable=Observable(slug="Erel", name="Erel", units="MeV"),
        computed=_computed(median=0.331, N_models=40),
        run_stamp=run_stamp,
        observable_path_slug="Erel_vs_16C_gs",
        reference=Reference(nucleus="16C", alias="gs"),
        artifacts={"histogram_plot": "histogram.png"},
        available=["histogram"],
    )


def _place(
    library: Path,
    record: ResultRecord,
    *,
    files: dict | None = None,
    path_id: str | None = None,
) -> Path:
    """Write a bundle at ``bundles/<id>/``, plus any extra artifact files."""
    result_id = path_id if path_id is not None else record.id
    bundle = library / "bundles" / Path(*result_id.split("/"))
    bundle.mkdir(parents=True, exist_ok=True)
    record.write_json(bundle / "result.json")
    extras = files if files is not None else {
        "histogram.png": _PLOT,
        "histogram.parquet": b"pq",
        "histogram.csv": _CSV,
    }
    for name, content in extras.items():
        payload = content if isinstance(content, bytes) else str(content).encode("utf-8")
        (bundle / name).write_bytes(payload)
    return bundle


@pytest.fixture
def library(tmp_path: Path) -> Path:
    """A small library: two 6He variants (one curated) plus a 16C transition."""
    root = tmp_path / "library"
    root.mkdir()
    first = _he6_record("2026-08-18_a")
    second = _he6_record("2026-08-18_b", computed=_computed(median=2.650, N_models=8))
    _place(root, first)
    _place(root, second)
    _place(root, _c16_transition(), files={"histogram.png": _PLOT})
    root.joinpath("annotations.toml").write_text(
        f'''["{first.id}"]
status  = "working"
title   = "6He Erel (0+, T=1)(2)"
tags    = ["hOmega-scan"]
note    = "Own sample weights."
outcome = "Promising."
''',
        encoding="utf-8",
    )
    return root


# ---------------------------------------------------------------------------
# Annotations
# ---------------------------------------------------------------------------

class TestAnnotations:
    def test_missing_file_is_empty(self, tmp_path):
        assert load_annotations(tmp_path) == {}

    def test_malformed_file_does_not_abort(self, tmp_path):
        (tmp_path / "annotations.toml").write_text("this is = not { toml", encoding="utf-8")
        assert load_annotations(tmp_path) == {}

    def test_overlay_wins_over_record_status(self):
        merged = merge_annotation(
            {"status": "probing"},
            {"status": "published", "note": "in the paper", "tags": ["v1"]},
        )
        assert merged["status"] == "published"
        assert merged["note"] == "in the paper"
        assert merged["tags"] == ["v1"]

    def test_lint_reports_orphan_unknown_key_and_bad_status(self):
        overlay = {
            "missing/id": {"status": "done", "typo_note": "oops"},
        }
        problems = lint_annotations(overlay, known_ids={"real/id"})
        joined = "\n".join(problems)
        assert "does not match any stored result" in joined
        assert "unknown annotation key" in joined
        assert "status 'done'" in joined


# ---------------------------------------------------------------------------
# Migrations
# ---------------------------------------------------------------------------

class TestMigrations:
    def test_current_records_need_no_upgrade(self):
        record = json.loads(_he6_record("stamp").to_json())
        upgraded, applied = apply_migrations(record)
        assert applied == []
        assert upgraded["schema_version"] == 2
        assert latest_version() == 2

    def test_v1_four_segment_id_gains_selection_folder(self):
        record = json.loads(_he6_record("2026-08-18_15-02-11").to_json())
        variant_hash = record["variant"]["variant_hash"]
        record["id"] = f"6He/0p-T1-n2/Erel/2026-08-18_15-02-11_{variant_hash}"
        record["schema_version"] = 1
        record["variant"].pop("cohort_hash", None)
        upgraded, applied = apply_migrations(record)
        assert applied == ["v1->v2"]
        assert upgraded["schema_version"] == 2
        parts = upgraded["id"].split("/")
        assert len(parts) == 5
        assert parts[4].startswith("final_")
        assert parts[4].endswith(variant_hash)
        assert upgraded["variant"]["cohort_hash"]
        assert parts[3].endswith(upgraded["variant"]["cohort_hash"])

    def test_registered_migration_runs_on_read(self):
        from results_library import migrations

        original = dict(migrations.MIGRATIONS)
        try:
            migrations.MIGRATIONS.clear()

            @migrations.register(1)
            def _v1_to_v2(record):
                computed = dict(record.get("computed") or {})
                computed.setdefault("uncertainty_method", "IQR")
                record["computed"] = computed
                return record

            record = {"schema_version": 1, "id": "x", "computed": {}}
            upgraded, applied = apply_migrations(record)
            assert applied == ["v1->v2"]
            assert upgraded["schema_version"] == 2
            assert upgraded["computed"]["uncertainty_method"] == "IQR"
        finally:
            migrations.MIGRATIONS.clear()
            migrations.MIGRATIONS.update(original)

    def test_future_schema_is_left_untouched(self):
        record = {"schema_version": 99, "id": "future"}
        upgraded, applied = apply_migrations(record)
        assert applied == []
        assert upgraded["schema_version"] == 99


# ---------------------------------------------------------------------------
# Catalog
# ---------------------------------------------------------------------------

class TestCatalog:
    def test_empty_library_writes_both_catalog_forms(self, tmp_path):
        library = tmp_path / "empty"
        library.mkdir()
        frame, scan = build_catalog(library)
        assert frame.empty
        assert scan.records == []
        paths = write_catalog(frame, library)
        assert paths["parquet"].exists()
        loaded = load_catalog(library)
        assert loaded.empty
        with sqlite3.connect(paths["sqlite"]) as connection:
            tables = {
                row[0]
                for row in connection.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'"
                )
            }
        assert "results" in tables

    def test_scan_skips_unreadable_json(self, tmp_path):
        bundle = tmp_path / "bundles" / "bad"
        bundle.mkdir(parents=True)
        (bundle / "result.json").write_text("{not json", encoding="utf-8")
        scan = scan_bundles(tmp_path)
        assert scan.records == []
        assert scan.unreadable

    def test_flatten_merges_annotations_and_labels(self, library):
        frame, scan = build_catalog(library)
        assert len(scan.records) == 3
        assert len(frame) == 3

        working = frame[frame["status"] == "working"].iloc[0]
        probing = frame[frame["status"] == "probing"]
        assert working["note"] == "Own sample weights."
        assert working["outcome"] == "Promising."
        assert working["tags"] == "hOmega-scan"
        assert working["nucleus"] == "6He"
        assert working["state_slug"] == "0p-T1-n2"
        assert len(str(working["id"]).split("/")) == 5
        assert str(working["selection_stamp"]).startswith("final_")
        assert "(2)" in working["state_label"]  # n>1 is shown
        assert working["median"] == pytest.approx(2.639)
        assert "2.639" in working["value_label"]
        assert working["artifact_histogram_plot"].endswith("histogram.png")
        assert probing["status"].eq("probing").all()
        assert working["observable_label"] == "E\u1d63\u2091\u2097"
        assert working["potential"] == "Daejeon16"
        assert working["nmax_range"] == "[2, 6]"
        assert working["setup_label"].startswith("Daejeon16")
        assert "\u0127\u03a9" in working["setup_label"]

        transition = frame[frame["nucleus"] == "16C"].iloc[0]
        assert transition["subject_kind"] == "transition"
        assert transition["direction"] == "down"
        assert "\u2192" in transition["state_label"]

    def test_shared_source_paths_and_provenance_check_flatten(self, tmp_path):
        library = tmp_path / "sources"
        library.mkdir()
        record = _he6_record(
            "src",
            provenance=Provenance(
                source_workbook="6He.xlsx",
                source_sheet="0p2",
                source_file=StoredFile(
                    name="raw.xlsx",
                    sha256="ab" * 32,
                    path="sources/abcd1234abcd1234/raw.xlsx",
                ),
                prepared_file=StoredFile(
                    name="long.xlsx",
                    sha256="cd" * 32,
                    path="sources/cdef5678cdef5678/long.xlsx",
                ),
                provenance_check={
                    "verdict": "pass",
                    "detail": "4/4 rows",
                    "n_rows": 4,
                    "n_rows_matched": 4,
                },
            ),
        )
        _place(library, record, files={})
        frame, _ = build_catalog(library)
        row = frame.iloc[0]
        assert row["source_file_path"] == "sources/abcd1234abcd1234/raw.xlsx"
        assert row["prepared_file_name"] == "long.xlsx"
        assert row["source_files_name"] == "raw.xlsx"
        assert row["provenance_check_verdict"] == "pass"
        assert row["provenance_check_n_rows_matched"] == 4

    def test_rebuild_is_from_scratch(self, library):
        frame, _ = build_catalog(library)
        write_catalog(frame, library)
        first = load_catalog(library)
        write_catalog(frame, library)
        second = load_catalog(library)
        pd.testing.assert_frame_equal(first, second)

    def test_sqlite_is_queryable(self, library):
        frame, _ = build_catalog(library)
        paths = write_catalog(frame, library)
        with sqlite3.connect(paths["sqlite"]) as connection:
            rows = connection.execute(
                "SELECT nucleus, COUNT(*) FROM results GROUP BY nucleus ORDER BY nucleus"
            ).fetchall()
        assert rows == [("16C", 1), ("6He", 2)]

    def test_load_catalog_without_build_explains_how(self, tmp_path):
        with pytest.raises(FileNotFoundError, match="Build it first"):
            load_catalog(tmp_path)


# ---------------------------------------------------------------------------
# Lint
# ---------------------------------------------------------------------------

class TestLint:
    def test_clean_library_has_no_errors(self, library):
        scan = scan_bundles(library)
        issues = lint_library(library, scan)
        errors = [i for i in issues if i.severity == "error"]
        assert errors == []

    def test_duplicate_ids(self, tmp_path):
        library = tmp_path / "dup"
        library.mkdir()
        record = _he6_record("same")
        _place(library, record)
        _place(library, record, path_id="other/copy/of/same", files={})
        issues = lint_library(library, scan_bundles(library))
        kinds = {i.kind for i in issues}
        assert "duplicate-id" in kinds

    def test_missing_artifact(self, tmp_path):
        library = tmp_path / "missing"
        library.mkdir()
        record = _he6_record("gone")
        _place(library, record, files={})  # declare artifacts, write none
        issues = lint_library(library, scan_bundles(library))
        kinds = {i.kind for i in issues}
        assert "missing-artifact" in kinds

    def test_conflicting_isospin(self, tmp_path):
        library = tmp_path / "t"
        library.mkdir()
        a = _he6_record("t-a")
        # Same (J, pi, n), different T. T2=4 is T=2, still allowed for 6He.
        b = ResultRecord.build(
            subject=Subject.single(
                "6He", StateRecord(J2=0, parity="+", T2=4, index=2)
            ),
            observable=Observable(slug="Erel", name="Erel", units="MeV"),
            computed=_computed(),
            run_stamp="t-b",
            artifacts={},
            available=[],
        )
        _place(library, a, files={})
        _place(library, b, files={})
        issues = lint_library(library, scan_bundles(library))
        kinds = {i.kind for i in issues}
        assert "conflicting-isospin" in kinds

    def test_path_mismatch_and_orphan_annotation(self, tmp_path):
        library = tmp_path / "mismatch"
        library.mkdir()
        record = _he6_record("here")
        _place(library, record, path_id="not/the/id/path", files={})
        library.joinpath("annotations.toml").write_text(
            '["ghost/id"]\nnote = "orphan"\n',
            encoding="utf-8",
        )
        issues = lint_library(library, scan_bundles(library))
        kinds = {i.kind for i in issues}
        assert "path-mismatch" in kinds
        assert "annotation" in kinds

    def test_duplicate_ids_do_not_abort_catalog_write(self, tmp_path):
        library = tmp_path / "dupwrite"
        library.mkdir()
        record = _he6_record("same")
        _place(library, record)
        _place(library, record, path_id="other/copy/of/same", files={})
        frame, _ = build_catalog(library)
        paths = write_catalog(frame, library)
        assert paths["sqlite"].exists()
        with sqlite3.connect(paths["sqlite"]) as connection:
            count = connection.execute("SELECT COUNT(*) FROM results").fetchone()[0]
        assert count == 2


# ---------------------------------------------------------------------------
# LaTeX / TSV / CSV
# ---------------------------------------------------------------------------

class TestLatex:
    def test_booktabs_uses_catalog_values_not_html(self):
        rows = [
            {
                "state_latex": "$(0^{+},\\ T{=}1)(2)$",
                "state_label": "(0⁺, T=1)(2)",
                "observable": "Erel",
                "observable_latex": "$E_{rel}$",
                "value_latex": "2.639^{+0.002}_{-0.002}",
                "value_label": "2.639 +0.002 / -0.002",
                "N_models": 12.0,
                "Nmax_final": 8.0,
            }
        ]
        tex = latex_table(rows)
        assert "\\begin{tabular}" in tex
        assert "\\toprule" in tex
        assert "$(0^{+},\\ T{=}1)(2)$" in tex
        assert "$2.639^{+0.002}_{-0.002}$" in tex
        assert "$E_{rel}$" in tex
        assert "12" in tex

        tsv = tsv_table(rows)
        assert "Erel" in tsv
        assert "2.639 +0.002 / -0.002" in tsv
        assert "\\toprule" not in tsv

        csv_text = csv_table(rows)
        assert csv_text.splitlines()[0].startswith("Date")
        assert "State" in csv_text.splitlines()[0]
        assert "Erel" in csv_text
        assert "Potential" in tex
        assert "$N_{\\max}$" in tex

    def test_latex_escapes_plain_text_but_not_math(self):
        rows = [{"observable": "A_b", "state_latex": "$(2^{+})$", "value_latex": "1"}]
        tex = latex_table(rows, columns=("state", "observable", "value"))
        assert "A\\_b" in tex
        assert "$(2^{+})$" in tex


# ---------------------------------------------------------------------------
# Site
# ---------------------------------------------------------------------------

class TestSite:
    def test_build_writes_index_nucleus_and_detail_pages(self, library):
        pytest.importorskip("jinja2")
        from results_library.views.site import build_site

        frame, _ = build_catalog(library)
        site = build_site(library, frame)

        index = (site / "index.html").read_text(encoding="utf-8")
        assert "hide probing" in index
        # Fixture has a curated (working) 6He result, so probing stays hidden.
        assert 'id="hide-probing" checked' in index
        assert "Copy LaTeX" in index
        assert "Copy TSV" in index
        assert (site / "assets" / "app.js").exists()
        assert (site / "exports" / "all.tex").exists()
        assert (site / "exports" / "all.csv").exists()

        readme = (library / "README.txt").read_bytes()
        assert readme == (Path(__file__).parent.parent / "results_library" / "views" / "library_readme.txt").read_bytes()
        assert b"Full Disk Access" in readme
        assert b"index.html" in readme

        root_index = (library / "index.html").read_text(encoding="utf-8")
        assert 'href="site/assets/style.css"' in root_index
        assert 'src="site/assets/app.js"' in root_index
        assert 'href="site/nucleus/6He.html"' in root_index
        assert 'href="site/exports/all.csv"' in root_index
        assert 'href="site/results/' in root_index
        assert 'href="assets/style.css"' in index
        assert 'href="site/assets/style.css"' not in index

        assert (site / "nucleus" / "6He.html").exists()
        assert (site / "nucleus" / "16C.html").exists()
        assert index.find("nucleus/6He.html") < index.find("nucleus/16C.html")

        working = frame[frame["status"] == "working"].iloc[0]
        page = site / "results" / (working["id"].replace("/", "__") + ".html")
        html = page.read_text(encoding="utf-8")
        assert "Own sample weights." in html
        assert "Promising." in html
        # From site/results/*.html the library root is two levels up, not three.
        assert 'src="../../bundles/' in html
        assert "../../../bundles/" not in html
        assert "E\u1d63\u2091\u2097" in html
        assert "\u0127\u03a9 aggregation" in html
        assert "Fingerprint of the setup" in html
        assert "help-mark" in html
        assert "Nmax at which the extrapolated value is read off" in html
        assert "Daejeon16" in html

    def test_detail_page_shows_selected_data_plot_first(self, tmp_path):
        pytest.importorskip("jinja2")
        from results_library.views.site import build_site

        library = tmp_path / "plots-order"
        library.mkdir()
        record = _he6_record(
            "2026-08-18_plots",
            artifacts={
                "histogram_plot": "histogram.png",
                "ensemble_curves_plot": "ensemble_curves.png",
                "selected_data_plot": "selected_data_Erel.jpg",
            },
            available=["histogram", "ensemble_curves", "selected_data"],
        )
        _place(
            library,
            record,
            files={
                "histogram.png": _PLOT,
                "ensemble_curves.png": _PLOT,
                "selected_data_Erel.jpg": _PLOT,
            },
        )
        frame, _ = build_catalog(library)
        site = build_site(library, frame)
        page = (
            site / "results" / (record.id.replace("/", "__") + ".html")
        ).read_text(encoding="utf-8")
        selected = page.find("Selected data")
        histogram = page.find("Histogram")
        ensemble = page.find("Ensemble curves")
        assert selected != -1
        assert histogram != -1
        assert ensemble != -1
        assert selected < histogram < ensemble
        assert "selected_data_Erel.jpg" in page

    def test_date_column_sorts_newest_first(self, tmp_path):
        pytest.importorskip("jinja2")
        from results_library.views.site import build_site

        library = tmp_path / "dates"
        library.mkdir()
        older = _he6_record(
            "2026-08-18_09-59-21_aaaaaaaa",
            variant=Variant(
                selection_set="final",
                bounds={"Nmax": [2, 6], "hOmega": [13, 40]},
                potential="Daejeon16",
            ),
        )
        newer = _he6_record(
            "2026-08-25_11-10-27_bbbbbbbb",
            computed=_computed(median=3.0, N_models=8),
            variant=Variant(
                selection_set="final",
                bounds={"Nmax": [2, 6], "hOmega": [13, 40]},
                potential="Daejeon16",
            ),
        )
        _place(library, older, files={})
        _place(library, newer, files={})
        frame, _ = build_catalog(library)
        labels = list(frame["run_date_label"])
        assert labels[0] == "25 Aug 2026"
        assert labels[1] == "18 Aug 2026"
        site = build_site(library, frame)
        index = (site / "index.html").read_text(encoding="utf-8")
        assert "25 Aug 2026" in index
        assert index.index("25 Aug 2026") < index.index("18 Aug 2026")
        assert "sorted-desc" in index
        assert 'data-value="2026-08-25T11:10:27"' in index

    def test_math_is_typeset_from_stored_latex(self, library):
        pytest.importorskip("jinja2")
        from results_library.views.site import build_site

        frame, _ = build_catalog(library)
        site = build_site(library, frame)

        # Vendored, so a collaborator opening the folder needs nothing installed.
        assert (site / "assets" / "katex" / "katex.min.js").exists()
        assert (site / "assets" / "katex" / "katex.min.css").exists()
        assert (site / "assets" / "katex" / "fonts" / "KaTeX_Main-Regular.woff2").exists()
        css = (site / "assets" / "katex" / "katex.min.css").read_text(encoding="utf-8")
        assert "0.18.4" in css
        assert (site / "assets" / "InterVariable-Italic.woff2").exists()

        index = (site / "index.html").read_text(encoding="utf-8")
        assert 'class="tex"' in index
        assert 'data-tex="E_{rel}"' in index  # $...$ belongs to the exports
        # The Unicode form stays in the span as the no-script fallback.
        assert "E\u1d63\u2091\u2097" in index
        # Copy LaTeX is untouched: still math mode, still from catalog values.
        assert "$E_{rel}$" in index
        assert "$2.639^{+0.002}_{-0.002}$" in index

        working = frame[frame["status"] == "working"].iloc[0]
        page = (site / "results" / (working["id"].replace("/", "__") + ".html")).read_text(
            encoding="utf-8"
        )
        assert 'data-tex="2.639^{+0.002}_{-0.002}"' in page
        assert 'data-tex="\\hbar\\Omega\\ \\text{aggregation}"' in page

    def test_empty_catalog_still_renders_index(self, tmp_path):
        pytest.importorskip("jinja2")
        from results_library.views.site import build_site

        library = tmp_path / "empty"
        library.mkdir()
        site = build_site(library, pd.DataFrame())
        html = (site / "index.html").read_text(encoding="utf-8")
        assert "No results yet" in html

    def test_all_probing_library_shows_rows_by_default(self, tmp_path):
        pytest.importorskip("jinja2")
        from results_library.views.site import build_site

        library = tmp_path / "probing-only"
        library.mkdir()
        _place(library, _he6_record("only"))
        frame, _ = build_catalog(library)
        site = build_site(library, frame)
        html = (site / "index.html").read_text(encoding="utf-8")
        assert 'id="hide-probing"' in html
        assert 'id="hide-probing" checked' not in html
        assert "(0⁺, T=1)(2)" in html or "(0" in html
        assert "E\u1d63\u2091\u2097" in html

    def test_multiple_source_files_listed_on_detail_page(self, tmp_path):
        pytest.importorskip("jinja2")
        from results_library.views.site import build_site

        library = tmp_path / "multi-source"
        library.mkdir()
        record = _he6_record(
            "src",
            provenance=Provenance(
                source_file=StoredFile(
                    name="extract.xlsx",
                    sha256="ee" * 32,
                    path="sources/eeeeeeeeeeeeeeee/extract.xlsx",
                ),
                source_files=[
                    StoredFile(
                        name="dump_a.xlsx",
                        sha256="aa" * 32,
                        path="sources/aaaaaaaaaaaaaaaa/dump_a.xlsx",
                    ),
                    StoredFile(
                        name="dump_b.xlsx",
                        sha256="bb" * 32,
                        path="sources/bbbbbbbbbbbbbbbb/dump_b.xlsx",
                    ),
                ],
                prepared_file=StoredFile(
                    name="long.xlsx",
                    sha256="cc" * 32,
                    path="sources/cccccccccccccccc/long.xlsx",
                ),
            ),
        )
        _place(library, record, files={})
        frame, _ = build_catalog(library)
        row = frame.iloc[0]
        assert "dump_a.xlsx" in row["source_files_name"]
        assert "dump_b.xlsx" in row["source_files_name"]
        site = build_site(library, frame)
        html = (site / "results" / (row["id"].replace("/", "__") + ".html")).read_text(
            encoding="utf-8"
        )
        assert "dump_a.xlsx" in html
        assert "dump_b.xlsx" in html
        assert "original calculation file" in html
        assert "extract.xlsx" in html
        assert "extracted pivot" in html
        assert "long.xlsx" in html

    def test_selection_set_tabs_default_to_final(self, tmp_path):
        pytest.importorskip("jinja2")
        from results_library.views.site import build_site

        library = tmp_path / "tabs"
        library.mkdir()
        bounds = {"Nmax": [2, 6], "hOmega": [13, 40]}
        provenance = Provenance(run_dir="run-one", source_sheet="0p2")
        _place(
            library,
            _he6_record(
                "2026-08-18_final",
                variant=Variant(
                    selection_set="final", bounds=bounds, potential="Daejeon16"
                ),
                provenance=provenance,
            ),
        )
        _place(
            library,
            _he6_record(
                "2026-08-18_all",
                variant=Variant(
                    selection_set="all_models", bounds=bounds, potential="Daejeon16"
                ),
                provenance=provenance,
            ),
        )
        frame, _ = build_catalog(library)
        site = build_site(library, frame)
        index = (site / "index.html").read_text(encoding="utf-8")
        assert "data-selection-tabs" not in index
        assert 'data-selection="all_models"' not in index
        assert 'data-selection="final"' in index
        assert 'data-selection-tab="all_models"' not in index
        all_csv = (site / "exports" / "all.csv").read_text(encoding="utf-8")
        nucleus_csv = (site / "exports" / "nucleus-6He.csv").read_text(encoding="utf-8")
        assert all_csv.count("\n") == 2  # header + final row
        assert nucleus_csv.count("\n") == 3  # header + both selection sets

        nucleus = (site / "nucleus" / "6He.html").read_text(encoding="utf-8")
        assert "data-selection-tabs" in nucleus
        assert 'data-selection-tab="all_models"' in nucleus

        final_row = frame[frame["selection_set"] == "final"].iloc[0]
        page = (
            site / "results" / (final_row["id"].replace("/", "__") + ".html")
        ).read_text(encoding="utf-8")
        assert "all_models" in page
        assert ">final<" in page

    def test_nmax_windows_section_groups_training_ranges(self, tmp_path):
        pytest.importorskip("jinja2")
        from results_library.views.site import build_site

        library = tmp_path / "nmax"
        library.mkdir()
        shared = dict(
            selection_set="final",
            potential="Daejeon16",
            filter_criteria=["all_passed"],
        )
        _place(
            library,
            _he6_record(
                "2026-08-18_n6",
                variant=Variant(bounds={"Nmax": [2, 6], "hOmega": [13, 40]}, **shared),
            ),
        )
        _place(
            library,
            _he6_record(
                "2026-08-18_n18",
                variant=Variant(bounds={"Nmax": [2, 18], "hOmega": [13, 40]}, **shared),
            ),
        )
        frame, _ = build_catalog(library)
        site = build_site(library, frame)
        index = (site / "index.html").read_text(encoding="utf-8")
        assert "[2, 6]" in index
        assert "[2, 18]" in index
        page = (
            site / "results" / (frame.iloc[0]["id"].replace("/", "__") + ".html")
        ).read_text(encoding="utf-8")
        assert "Other Nmax windows" in page
        assert "[2, 6]" in page
        assert "[2, 18]" in page

    def test_comparisons_page_and_cross_links(self, library):
        pytest.importorskip("jinja2")
        from results_library.views.site import build_site

        frame, _ = build_catalog(library)
        working = frame[frame["status"] == "working"].iloc[0]
        (library / "comparisons").mkdir()
        (library / "comparisons" / "triplet.png").write_bytes(_PLOT)
        (library / "comparisons.toml").write_text(
            f'''[[figure]]
title = "Isospin triplet"
file = "comparisons/triplet.png"
caption = "A hand-made comparison."
results = ["{working["id"]}"]
''',
            encoding="utf-8",
        )
        site = build_site(library, frame)
        comparisons = (site / "comparisons.html").read_text(encoding="utf-8")
        assert "Isospin triplet" in comparisons
        assert "A hand-made comparison." in comparisons
        assert "../comparisons/triplet.png" in comparisons
        index = (site / "index.html").read_text(encoding="utf-8")
        assert "Comparisons" in index
        nucleus = (site / "nucleus" / "6He.html").read_text(encoding="utf-8")
        assert "Isospin triplet" in nucleus
        detail = (
            site / "results" / (working["id"].replace("/", "__") + ".html")
        ).read_text(encoding="utf-8")
        assert "Isospin triplet" in detail


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

class TestExcel:
    def test_probing_omitted_unless_requested(self, library):
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook
        from results_library.views.excel import build_workbook

        frame, _ = build_catalog(library)
        path = build_workbook(library, frame, include_probing=False)
        workbook = load_workbook(path)
        assert "6He" in workbook.sheetnames
        he6 = workbook["6He"]
        # Header + the one curated row (the other 6He result is probing).
        assert he6.max_row == 2
        assert he6.cell(row=1, column=1).value == "Date"
        assert he6.cell(row=2, column=4).value == "Daejeon16"
        assert he6.cell(row=2, column=10).value == "[2, 6]"
        assert he6.cell(row=2, column=12).value == "working"
        assert he6.cell(row=2, column=13).value == "Own sample weights."

        # 16C transition is still probing, so it is not shown.
        assert "16C" not in workbook.sheetnames

        full = build_workbook(library, frame, include_probing=True)
        shown = load_workbook(full)
        assert "16C" in shown.sheetnames
        assert shown["6He"].max_row == 3
        # Periodic-table order: He (Z=2) before C (Z=6). "about" is first.
        assert shown.sheetnames.index("6He") < shown.sheetnames.index("16C")

    def test_hyperlinks_point_at_plot_and_detail_page(self, library):
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook
        from results_library.views.excel import build_workbook

        frame, _ = build_catalog(library)
        path = build_workbook(library, frame, include_probing=False)
        sheet = load_workbook(path)["6He"]
        plot_cell = sheet.cell(row=2, column=18)  # Plot
        details = sheet.cell(row=2, column=21)  # Details
        assert plot_cell.hyperlink is not None
        assert "histogram.png" in str(plot_cell.hyperlink.target)
        assert details.hyperlink is not None
        assert str(details.hyperlink.target).endswith(".html")

    def test_empty_probing_only_library_explains_itself(self, tmp_path):
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook
        from results_library.views.excel import build_workbook

        library = tmp_path / "probing-only"
        library.mkdir()
        _place(library, _he6_record("only"), files={})
        frame, _ = build_catalog(library)
        path = build_workbook(library, frame, include_probing=False)
        sheet = load_workbook(path).worksheets[1]  # after the 'about' sheet
        assert "exploratory" in str(sheet.cell(row=2, column=1).value).lower()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

class TestCLI:
    def test_build_writes_catalog_site_and_workbook(self, library):
        pytest.importorskip("jinja2")
        pytest.importorskip("openpyxl")
        assert cli_main(["build", "--library", str(library)]) == 0
        assert (library / "catalog.parquet").exists()
        assert (library / "catalog.db").exists()
        assert (library / "site" / "index.html").exists()
        assert (library / "index.html").exists()
        assert (library / "README.txt").exists()
        assert (library / "calculation_results.xlsx").exists()
        # The human layer is never overwritten.
        text = (library / "annotations.toml").read_text(encoding="utf-8")
        assert "Own sample weights." in text

    def test_lint_exits_nonzero_on_errors(self, tmp_path):
        library = tmp_path / "bad"
        library.mkdir()
        record = _he6_record("same")
        _place(library, record, files={})
        _place(library, record, path_id="other/copy/of/same", files={})
        assert cli_main(["lint", "--library", str(library)]) == 1

    def test_missing_library_path_exits_2(self, tmp_path, monkeypatch):
        monkeypatch.delenv("RESULTS_LIBRARY", raising=False)
        with pytest.raises(SystemExit) as caught:
            cli_main(["build"])
        assert caught.value.code == 2

    def test_config_supplies_library_path(self, library, tmp_path):
        pytest.importorskip("jinja2")
        config = tmp_path / "config.toml"
        escaped = str(library).replace("\\", "/")
        config.write_text(
            f'[results_library]\npath = "{escaped}"\n',
            encoding="utf-8",
        )
        assert cli_main(["catalog", "--config", str(config)]) == 0
        assert (library / "catalog.parquet").exists()

    def test_annotations_file_untouched_by_build(self, library):
        pytest.importorskip("jinja2")
        original = (library / "annotations.toml").read_bytes()
        assert cli_main(["build", "--library", str(library), "--no-excel"]) == 0
        assert (library / "annotations.toml").read_bytes() == original

    def test_migrate_selection_folders_rewrites_v1_layout(self, tmp_path):
        library = tmp_path / "legacy"
        library.mkdir()
        record = _he6_record("2026-08-18_15-02-11")
        data = json.loads(record.to_json())
        variant_hash = data["variant"]["variant_hash"]
        old_id = f"6He/0p-T1-n2/Erel/2026-08-18_15-02-11_{variant_hash}"
        data["id"] = old_id
        data["schema_version"] = 1
        data["variant"].pop("cohort_hash", None)
        bundle = library / "bundles" / Path(*old_id.split("/"))
        bundle.mkdir(parents=True)
        (bundle / "result.json").write_text(json.dumps(data), encoding="utf-8")
        (library / "annotations.toml").write_text(
            f'["{old_id}"]\nnote = "keep"\n',
            encoding="utf-8",
        )

        assert cli_main(["migrate-selection-folders", "--library", str(library)]) == 0
        remaining = list((library / "bundles").rglob("result.json"))
        assert len(remaining) == 1
        moved = json.loads(remaining[0].read_text(encoding="utf-8"))
        parts = moved["id"].split("/")
        assert len(parts) == 5
        assert parts[4].startswith("final_")
        assert remaining[0].parent.name.startswith("final_")
        assert remaining[0].parent.parent.name.endswith(moved["variant"]["cohort_hash"])
        assert old_id not in (library / "annotations.toml").read_text(encoding="utf-8")
        assert moved["id"] in (library / "annotations.toml").read_text(encoding="utf-8")

    def test_library_flag_works_before_or_after_subcommand(self, library):
        assert cli_main(["--library", str(library), "catalog"]) == 0
        assert cli_main(["catalog", "--library", str(library)]) == 0


# ---------------------------------------------------------------------------
# Vendor copy
# ---------------------------------------------------------------------------

class TestVendorSync:
    def test_vendor_schema_is_complete_and_importable(self):
        vendor = ROOT / "results_library" / "vendor" / "results_schema"
        for name in SCHEMA_FILES:
            assert (vendor / name).exists(), name
        from results_schema.models import SCHEMA_VERSION

        assert SCHEMA_VERSION == 2

    def test_vendor_matches_pipeline_when_source_is_available(self):
        source = source_dir()
        if source is None or not source.is_dir():
            sibling = ROOT.parent / "2026 rework" / "results_schema"
            source = sibling if sibling.is_dir() else None
        if source is None:
            pytest.skip("pipeline results_schema not checked out next to this repo")
        stale = sync_schema(check_only=True, source=source)
        assert stale == 0, "run: python -m results_library.sync_schema --source <pipeline>/results_schema"

    def test_sync_without_source_exits_2(self, monkeypatch):
        from results_library.sync_schema import main as sync_main

        monkeypatch.delenv("RESULTS_SCHEMA_SOURCE", raising=False)
        assert sync_main(["--check"]) == 2


# ---------------------------------------------------------------------------
# Cross-nucleus identity in the catalog
# ---------------------------------------------------------------------------

class TestCrossNucleus:
    def test_primary_subject_plus_reference(self, tmp_path):
        library = tmp_path / "xref"
        library.mkdir()
        record = _c17_relative()
        _place(library, record, files={"histogram.png": _PLOT})
        frame, _ = build_catalog(library)
        row = frame.iloc[0]
        assert row["nucleus"] == "17C"
        assert row["observable_slug"] == "Erel_vs_16C_gs"
        assert row["reference_nucleus"] == "16C"
        assert row["reference_alias"] == "gs"
        assert "--to--" not in row["id"]
