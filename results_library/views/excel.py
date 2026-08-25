"""
Excel view of the catalog.

Regenerates the familiar master-workbook layout that used to be maintained by
hand: one sheet per nucleus, rows grouped by state, showing the central value
with asymmetric uncertainty and the number of models. Every row carries real
hyperlinks to its plot, its data file, and its detail page.

This is a view like any other: it is rewritten from scratch on every build and is
never a source of truth. Notes come from ``annotations.toml`` and the numbers
come from the bundles, so nothing here has to be typed again.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd

from results_schema.labels import axis_unicode

logger = logging.getLogger(__name__)

WORKBOOK_FILENAME = "calculation_results.xlsx"
_MAX_SHEET_NAME = 31

#: (column header, catalog key)
_COLUMNS: tuple[tuple[str, str], ...] = (
    ("Date", "run_date_label"),
    ("State", "state_label"),
    ("Observable", "observable_label"),
    ("Potential", "potential"),
    ("Units", "units"),
    ("median - Q1", "err_low"),
    ("median", "median"),
    ("Q3 - median", "err_high"),
    ("N models", "N_models"),
    (f"{axis_unicode('Nmax')} [min, max]", "nmax_range"),
    (f"{axis_unicode('Nmax')} final", "Nmax_final"),
    ("Status", "status"),
    ("Note", "note"),
    ("Outcome", "outcome"),
    ("Tags", "tags"),
    ("Run", "run_stamp"),
    ("Variant", "variant_hash"),
)

_LINK_COLUMNS: tuple[tuple[str, str], ...] = (
    ("Plot", "artifact_histogram_plot"),
    ("Data", "artifact_histogram_data"),
    ("CSV", "artifact_histogram_data_csv"),
)


def _clean(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and value != value:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value


def _safe_sheet_name(name: str, used: set[str]) -> str:
    """Excel forbids several characters and caps names at 31 chars."""
    cleaned = str(name)
    for char in "[]:*?/\\":
        cleaned = cleaned.replace(char, "-")
    cleaned = cleaned[:_MAX_SHEET_NAME] or "sheet"

    candidate = cleaned
    suffix = 2
    while candidate in used:
        tail = f"_{suffix}"
        candidate = cleaned[: _MAX_SHEET_NAME - len(tail)] + tail
        suffix += 1
    used.add(candidate)
    return candidate


def build_workbook(
    library_root: Path,
    catalog: Optional[pd.DataFrame] = None,
    include_probing: bool = False,
) -> Path:
    """Write ``calculation_results.xlsx`` and return its path.

    Exploratory results are excluded by default: the workbook is the quick-look
    surface, and hundreds of probing rows would bury the handful that matter.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    library_root = Path(library_root)
    if catalog is None:
        from results_library.catalog import load_catalog

        catalog = load_catalog(library_root)

    frame = catalog
    if not include_probing and not frame.empty and "status" in frame:
        frame = frame[frame["status"] != "probing"]

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_names: set[str] = set()

    headers = [header for header, _ in _COLUMNS] + [header for header, _ in _LINK_COLUMNS] + ["Details"]

    if frame.empty:
        sheet = workbook.create_sheet(_safe_sheet_name("results", used_names))
        sheet.append(headers)
        _style_header(sheet, Font, Alignment)
        note = (
            "No results to show."
            if catalog.empty
            else "Only exploratory (probing) results are stored so far. "
            "Promote one in annotations.toml to see it here."
        )
        sheet.cell(row=2, column=1, value=note)
    else:
        for nucleus, rows in frame.groupby("nucleus", sort=True):
            sheet = workbook.create_sheet(_safe_sheet_name(str(nucleus), used_names))
            sheet.append(headers)
            _style_header(sheet, Font, Alignment)

            sort_cols = [
                c
                for c in ("run_datetime", "state_slug", "observable", "id")
                if c in rows.columns
            ]
            ascending = [c != "run_datetime" for c in sort_cols]
            ordered = (
                rows.sort_values(
                    sort_cols, ascending=ascending, na_position="last"
                )
                if sort_cols
                else rows
            )
            for _, record in ordered.iterrows():
                _append_row(sheet, record, library_root, get_column_letter)

            _autosize(sheet, get_column_letter)
            sheet.freeze_panes = "A2"

    summary = workbook.create_sheet(_safe_sheet_name("about", used_names), 0)
    _write_about(summary, catalog, frame, Font)

    path = library_root / WORKBOOK_FILENAME
    try:
        workbook.save(path)
    except PermissionError:
        # The workbook is very likely open in Excel; say so rather than dumping a
        # traceback, since the catalog itself is already safely written.
        raise PermissionError(
            f"Cannot write {path}. Close it in Excel and rebuild."
        ) from None
    return path


def _style_header(sheet, Font, Alignment) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _append_row(sheet, record, library_root: Path, get_column_letter) -> None:
    values = []
    for _, key in _COLUMNS:
        value = _clean(record.get(key))
        if key == "observable_label" and not value:
            value = _clean(record.get("observable"))
        values.append(value)
    sheet.append(values)
    row_index = sheet.max_row

    column = len(_COLUMNS) + 1
    for label, key in _LINK_COLUMNS:
        target = _clean(record.get(key))
        cell = sheet.cell(row=row_index, column=column)
        if target:
            # Relative to the workbook, which sits at the library root, so links
            # keep working wherever the library is synced.
            cell.value = label
            cell.hyperlink = str(target).replace("/", "\\")
            cell.style = "Hyperlink"
        column += 1

    result_id = _clean(record.get("id"))
    cell = sheet.cell(row=row_index, column=column)
    if result_id:
        page = str(result_id).replace("/", "__") + ".html"
        cell.value = "open"
        cell.hyperlink = f"site\\results\\{page}"
        cell.style = "Hyperlink"

    for offset, (_, key) in enumerate(_COLUMNS, start=1):
        if key in ("median", "err_low", "err_high"):
            sheet.cell(row=row_index, column=offset).number_format = "0.000"


def _autosize(sheet, get_column_letter, minimum: int = 8, maximum: int = 46) -> None:
    for cells in sheet.columns:
        lengths = [len(str(c.value)) for c in cells if c.value is not None]
        if not lengths:
            continue
        width = min(max(max(lengths) + 2, minimum), maximum)
        sheet.column_dimensions[get_column_letter(cells[0].column)].width = width


def _write_about(sheet, catalog: pd.DataFrame, shown: pd.DataFrame, Font) -> None:
    sheet["A1"] = "Calculation results"
    sheet["A1"].font = Font(bold=True, size=14)

    lines: List[tuple[str, Any]] = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Results stored", 0 if catalog.empty else len(catalog)),
        ("Results shown here", 0 if shown.empty else len(shown)),
    ]
    if not catalog.empty and "status" in catalog:
        for status, count in catalog["status"].value_counts().items():
            lines.append((f"  status: {status}", int(count)))

    lines.extend(
        [
            ("", ""),
            ("This workbook is GENERATED", "Do not edit; changes are lost on rebuild."),
            ("Numbers come from", "bundles/<id>/result.json"),
            ("Notes come from", "annotations.toml (hand-edited, never overwritten)"),
            ("Rebuild with", "python -m results_library.cli build"),
            ("", ""),
            ("Probing results", "hidden by default; pass --include-probing to show them"),
        ]
    )

    for offset, (label, value) in enumerate(lines, start=3):
        sheet.cell(row=offset, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=offset, column=2, value=value)

    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 62
